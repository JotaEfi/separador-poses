import os
import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askdirectory, askopenfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule


def carregar_planilha(caminho_planilha):
    df = pd.read_excel(caminho_planilha)
    df.reset_index(drop=True, inplace=True)  # Reseta o índice
    if "Unnamed: 0" in df.columns:
        df.rename(columns={"Unnamed: 0": "ID"}, inplace=True)
        df["ID"] = range(1, len(df) + 1)  # Atribui valores crescentes à coluna ID
    return df


def listar_pastas(pasta):
    return [
        item for item in os.listdir(pasta) if os.path.isdir(os.path.join(pasta, item))
    ]


def contar_fotos_por_pose(pasta_fotos, nomes_poses):
    contagem = {pose: 0 for pose in nomes_poses}
    for pose in nomes_poses:
        caminho_pose = os.path.join(pasta_fotos, pose)
        if os.path.exists(caminho_pose):
            contagem[pose] = len(
                os.listdir(caminho_pose)
            )  # Conta os arquivos na pasta da pose
    return contagem


def contar_fotos_alunos(pasta_alunos, alunos, nomes_poses, pasta_poses):
    contagem_alunos = {}
    for aluno in alunos["Nome"]:  # Assume que a coluna com os nomes se chama "Nome"
        caminho_aluno = os.path.join(pasta_alunos, aluno)
        if os.path.isdir(caminho_aluno):
            contagem_alunos[aluno] = {
                pose: 0 for pose in nomes_poses
            }  # Inicializa contagem para poses
            fotos_aluno = os.listdir(caminho_aluno)  # Lista todas as fotos do aluno

            # Contar fotos totais escolhidas
            total_fotos = len(fotos_aluno)

            for foto in fotos_aluno:
                for pose in nomes_poses:
                    caminho_pose = os.path.join(
                        pasta_poses, pose
                    )  # Caminho da pasta da pose
                    if os.path.exists(caminho_pose) and foto in os.listdir(
                        caminho_pose
                    ):  # Verifica se a foto pertence à pose
                        contagem_alunos[aluno][pose] += 1

            # Adiciona a contagem total de fotos
            contagem_alunos[aluno]["Total"] = total_fotos

    return contagem_alunos


def gerar_planilha_resultado(alunos, contagem_fotos, contagem_alunos):
    df_resultado = pd.DataFrame(alunos)
    for aluno in contagem_alunos:
        for pose in contagem_fotos.keys():
            df_resultado.loc[df_resultado["Nome"] == aluno, pose] = contagem_alunos[
                aluno
            ].get(pose, 0)
        # Adiciona a coluna Total
        df_resultado.loc[df_resultado["Nome"] == aluno, "Total"] = contagem_alunos[
            aluno
        ]["Total"]
    return df_resultado


def adicionar_somas(df_resultado, caminho_arquivo):
    # Salva o dataframe em um arquivo Excel
    writer = pd.ExcelWriter(caminho_arquivo, engine="openpyxl")
    df_resultado.to_excel(writer, index=False)
    writer.close()  # Substituído writer.save() por writer.close()

    # Carregar o workbook gerado
    workbook = load_workbook(caminho_arquivo)
    worksheet = workbook.active

    # Adicionar fórmula de soma nas colunas de poses
    total_rows = worksheet.max_row
    total_columns = worksheet.max_column

    for col in range(4, total_columns):  # Começa na coluna 4 (D)
        coluna_letra = chr(64 + col)  # Converte número da coluna para letra
        formula = f"=SUM({coluna_letra}2:{coluna_letra}{total_rows})"
        worksheet[f"{coluna_letra}{total_rows + 1}"] = formula

    # Definir o intervalo das células de soma (linha de soma das poses)
    soma_range = f"D{total_rows + 1}:{chr(64 + total_columns)}{total_rows + 1}"

    # Formatação condicional com gradiente verde (mais escuro para o maior valor, mais claro para o menor valor)
    rule = ColorScaleRule(
        start_type="min",
        start_color="C6EFCE",  # Verde claro para o valor mínimo
        end_type="max",
        end_color="006100",  # Verde escuro para o valor máximo
    )
    worksheet.conditional_formatting.add(soma_range, rule)

    # Salvar o workbook com as fórmulas e formatação condicional aplicadas
    workbook.save(caminho_arquivo)


def selecionar_pasta():
    Tk().withdraw()  # Oculta a janela principal
    return askdirectory(title="Selecione a pasta com as fotos")


def selecionar_planilha():
    Tk().withdraw()  # Oculta a janela principal
    return askopenfilename(
        title="Selecione a planilha com os pacotes",
        filetypes=[("Excel files", "*.xlsx;*.xls")],
    )


def main():
    # Selecionar a pasta com as fotos das poses
    global pasta_fotos
    pasta_fotos = selecionar_pasta()  # Seleciona a pasta das poses
    if not pasta_fotos:
        print("Nenhuma pasta selecionada.")
        return

    # Selecionar a planilha com os pacotes
    caminho_planilha = selecionar_planilha()
    if not caminho_planilha:
        print("Nenhuma planilha selecionada.")
        return

    # Carregar a planilha com os pacotes
    alunos = carregar_planilha(caminho_planilha)

    # Listar pastas (poses) na pasta de fotos
    nomes_poses = listar_pastas(pasta_fotos)

    # Contar fotos por pose
    contagem_fotos = contar_fotos_por_pose(pasta_fotos, nomes_poses)

    # Selecionar a pasta com as fotos dos alunos
    pasta_alunos = selecionar_pasta()  # Permite selecionar a pasta dos alunos
    contagem_alunos = contar_fotos_alunos(pasta_alunos, alunos, nomes_poses)

    # Gerar a planilha de resultado
    resultado = gerar_planilha_resultado(alunos, contagem_fotos, contagem_alunos)

    # Salvar a nova planilha
    caminho_arquivo = "resultado_fotos.xlsx"
    resultado.to_excel(caminho_arquivo, index=False)

    # Adicionar somas e formatação condicional
    adicionar_somas(resultado, caminho_arquivo)


if __name__ == "__main__":
    main()
